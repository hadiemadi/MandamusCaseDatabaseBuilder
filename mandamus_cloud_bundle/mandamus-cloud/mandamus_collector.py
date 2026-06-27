#!/usr/bin/env python3
import os, sys, json, time, re
from datetime import datetime
import requests
TOKEN = os.environ.get("COURTLISTENER_TOKEN", "").strip()
if not TOKEN:
    sys.exit("ERROR: set COURTLISTENER_TOKEN env var.")
OUTDIR  = os.environ.get("OUTDIR", os.path.dirname(os.path.abspath(__file__)))
API     = "https://www.courtlistener.com/api/rest/v4"
HEADERS = {"Authorization": f"Token {TOKEN}"}
QUERIES = ['"administrative processing" visa mandamus','"221(g)" visa mandamus',
    '"unreasonable delay" consular immigrant visa','consular visa "writ of mandamus" delay',
    'TRAC factors "administrative processing" visa']
FILED_AFTER="2020-01-01"; FILED_BEFORE=None; PAGES_PER_QUERY=6
CAP_MIN,CAP_HR,CAP_DAY=5,50,125; MIN_THROTTLE=13
CKPT=os.path.join(OUTDIR,"mandamus_checkpoint.json")
RAW =os.path.join(OUTDIR,"mandamus_phase15_raw.json")
XLSX=os.path.join(OUTDIR,"mandamus_phase15_timing.xlsx")
class RateLimiter:
    def __init__(self,stamps): self.stamps=list(stamps)
    def _c(self,now,w): return sum(1 for t in self.stamps if now-t<w)
    def wait(self):
        while True:
            now=time.time(); self.stamps=[t for t in self.stamps if now-t<86400]
            waits=[MIN_THROTTLE-(now-self.stamps[-1]) if self.stamps else 0]
            for cap,win in ((CAP_MIN,60),(CAP_HR,3600),(CAP_DAY,86400)):
                if self._c(now,win)>=cap:
                    oldest=min(t for t in self.stamps if now-t<win)
                    waits.append((oldest+win)-now+1)
            w=max(waits+[0])
            if w<=0: self.stamps.append(time.time()); return
            if w>120:
                until=datetime.fromtimestamp(time.time()+w).strftime("%H:%M:%S")
                print(f"  [rate] cap reached, sleeping {int(w)}s (until ~{until}); checkpointed, safe to wait.",flush=True)
            time.sleep(min(w,300))
def api_get(rl,url,params=None):
    for _ in range(6):
        rl.wait()
        try: r=requests.get(url,headers=HEADERS,params=params,timeout=90)
        except requests.RequestException as e:
            print(f"  [net] {e}; retry 30s",flush=True); time.sleep(30); continue
        if r.status_code==429:
            b=int(r.headers.get("Retry-After","60"))+5; print(f"  [429] backoff {b}s",flush=True); time.sleep(b); continue
        if r.status_code>=500: time.sleep(20); continue
        return r
    return None
RX_MTD=re.compile(r"motions?\s+to\s+dismiss",re.I); RX_G=re.compile(r"grant",re.I)
RX_D=re.compile(r"den(y|ie|ying)",re.I); RX_O=re.compile(r"\b(order|memorandum|opinion|minute (order|entry)|ruling)\b",re.I)
RX_V=re.compile(r"(stipulat\w*\s+\w*\s*dismiss|voluntar\w*\s+dismiss|notice of (voluntary )?dismissal|rule\s*41)",re.I)
RX_P=re.compile(r"(judgment|grant\w*).{0,40}(plaintiff|petition|mandamus|writ)",re.I)
def parse_entries(entries):
    ev={"mtd_filed":None,"mtd_ruling":None,"mtd_outcome":None,"termination_type":None,"snippets":{}}
    for en in entries:
        d=en.get("description") or ""; dl=d.lower(); dt=en.get("date_filed")
        if not d: continue
        is_mtd=bool(RX_MTD.search(dl)); is_order=bool(RX_O.search(dl)) or "granting" in dl or "denying" in dl
        if is_mtd and is_order and (RX_G.search(dl) or RX_D.search(dl)):
            g,n=bool(RX_G.search(dl)),bool(RX_D.search(dl))
            ev["mtd_ruling"]=dt; ev["mtd_outcome"]="partial" if (g and n) else ("granted" if g else "denied")
            ev["snippets"]["mtd_ruling"]=d[:300]
        elif is_mtd and not is_order and ev["mtd_filed"] is None:
            ev["mtd_filed"]=dt; ev["snippets"]["mtd_filed"]=d[:300]
        if RX_V.search(dl):
            ev["termination_type"]="settled (voluntary/stipulated dismissal)"; ev["snippets"]["termination"]=d[:300]
        elif is_mtd and is_order and RX_G.search(dl) and not RX_D.search(dl):
            ev["termination_type"]="dismissed on MTD (government win)"; ev["snippets"].setdefault("termination",d[:300])
        elif RX_P.search(dl):
            ev["termination_type"]="plaintiff win (judgment/mandamus granted)"; ev["snippets"]["termination"]=d[:300]
    if ev["termination_type"] is None: ev["termination_type"]="other / unclassified"
    return ev
def dbtw(a,b):
    try: return (datetime.fromisoformat(b)-datetime.fromisoformat(a)).days
    except: return None
def load_ckpt():
    if os.path.exists(CKPT):
        return json.load(open(CKPT))
    return {"phase":"collect","stamps":[],"dockets":{},"done_ids":[],"records":{}}
def save_ckpt(c):
    json.dump(c,open(CKPT+".tmp","w")); os.replace(CKPT+".tmp",CKPT)
def collect(c,rl):
    print("PHASE: collecting dockets",flush=True)
    for q in QUERIES:
        print("  query:",q,flush=True)
        params={"q":q,"type":"d","filed_after":FILED_AFTER,"order_by":"dateFiled desc","page_size":20}
        if FILED_BEFORE: params["filed_before"]=FILED_BEFORE
        url=f"{API}/search/"
        for _ in range(PAGES_PER_QUERY):
            r=api_get(rl,url,params); c["stamps"]=rl.stamps
            if not r or r.status_code!=200: break
            data=r.json()
            for it in data.get("results",[]):
                did=it.get("docket_id")
                if did:
                    did=str(did)
                    if did not in c["dockets"]:
                        c["dockets"][did]={"docket_id":did,"case_name":it.get("caseName"),
                          "court":it.get("court"),"date_filed":it.get("dateFiled"),
                          "date_terminated":it.get("dateTerminated"),
                          "url":"https://www.courtlistener.com"+(it.get("absolute_url") or "")}
            save_ckpt(c); nxt=data.get("next")
            if not nxt: break
            url,params=nxt,None
    c["phase"]="mine"; save_ckpt(c)
    print("  total",len(c["dockets"]),"terminated",sum(1 for d in c["dockets"].values() if d["date_terminated"]),flush=True)
def mine(c,rl):
    print("PHASE: mining entries",flush=True)
    targets=[d for d in c["dockets"].values() if d["date_terminated"]]; done=set(c["done_ids"])
    for i,dk in enumerate(targets,1):
        did=dk["docket_id"]
        if did in done: continue
        entries=[]; url=f"{API}/docket-entries/"; params={"docket":did,"order_by":"date_filed","page_size":100}
        ok=True
        while url:
            r=api_get(rl,url,params); c["stamps"]=rl.stamps
            if not r or r.status_code!=200: ok=False; break
            data=r.json()
            for en in data.get("results",[]):
                entries.append({"date_filed":en.get("date_filed"),"description":en.get("description")})
            url=data.get("next"); params=None
        ev=parse_entries(entries); rec=dict(dk); rec.update(ev)
        rec["n_entries"]=len(entries); rec["entries_complete"]=ok and len(entries)>0
        rec["pendency_days"]=dbtw(dk["date_filed"],dk["date_terminated"])
        rec["mtd_denied_to_close_days"]=dbtw(ev["mtd_ruling"],dk["date_terminated"]) if (ev["mtd_ruling"] and ev["mtd_outcome"]=="denied") else None
        c["records"][did]=rec; c["done_ids"].append(did); save_ckpt(c)
        print(f"  [{i}/{len(targets)}] {(dk['case_name'] or '')[:42]:42} entries={len(entries)} {ev['termination_type'][:22]}",flush=True)
    c["phase"]="report"; save_ckpt(c)
def pct(v,p):
    if not v: return None
    return v[min(len(v)-1,max(0,int(round(p/100*(len(v)-1)))))]
def report(c):
    print("PHASE: report",flush=True)
    recs=list(c["records"].values()); json.dump(recs,open(RAW,"w"),indent=2)
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment
    from collections import Counter
    wb=Workbook(); H=Font(bold=True,color="FFFFFF"); HF=PatternFill("solid",start_color="1F4E78")
    settled=[r for r in recs if r["termination_type"].startswith("settled")]
    pa=sorted(r["pendency_days"] for r in recs if r["pendency_days"] is not None)
    ps=sorted(r["pendency_days"] for r in settled if r["pendency_days"] is not None)
    gp=sorted(r["mtd_denied_to_close_days"] for r in recs if r.get("mtd_denied_to_close_days") is not None)
    def row(lbl,v): return [lbl,len(v),v[0] if v else "-",pct(v,25),pct(v,50),pct(v,75),v[-1] if v else "-"]
    ws=wb.active; ws.title="Summary"; ws.append(["Metric","N","min","p25","median","p75","max"])
    ws.append(row("Filing->closure (all closed), days",pa))
    ws.append(row("Filing->settlement (voluntary dismissal), days",ps))
    ws.append(row("MTD denied->closure, days",gp)); ws.append([])
    ws.append(["Termination type","count"])
    for k,v in Counter(r["termination_type"] for r in recs).most_common(): ws.append([k,v])
    ws.append([]); ws.append(["MTD outcome","count"])
    for k,v in Counter(r["mtd_outcome"] for r in recs if r.get("mtd_outcome")).most_common(): ws.append([k,v])
    ws.append([]); ws.append(["CAVEAT","Survivorship bias: settled cases not in RECAP are invisible; 'closed' mixes settlement, MTD-grant, plaintiff win."])
    for cc in ws[1]: cc.font=H; cc.fill=HF
    ws.column_dimensions["A"].width=48
    cols=["docket_id","case_name","court","date_filed","date_terminated","pendency_days",
      "termination_type","mtd_filed","mtd_ruling","mtd_outcome","mtd_denied_to_close_days",
      "n_entries","entries_complete","url"]
    ws2=wb.create_sheet("Dockets"); ws2.append(cols)
    for r in sorted(recs,key=lambda x:x["date_filed"] or "",reverse=True): ws2.append([r.get(k) for k in cols])
    for cc in ws2[1]: cc.font=H; cc.fill=HF
    ws2.freeze_panes="A2"; ws2.auto_filter.ref=ws2.dimensions
    wb.save(XLSX); c["phase"]="done"; save_ckpt(c)
    print(f"DONE. Wrote:\n  {RAW}\n  {XLSX}\n  settled N={len(settled)}  MTD-denied gaps N={len(gp)}",flush=True)
def main():
    c=load_ckpt(); rl=RateLimiter(c.get("stamps",[]))
    if c["phase"]=="collect": collect(c,rl)
    if c["phase"]=="mine": mine(c,rl)
    if c["phase"] in ("report","done"): report(c)
    print("All phases complete.")
if __name__=="__main__": main()
